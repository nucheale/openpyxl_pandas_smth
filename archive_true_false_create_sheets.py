import openpyxl as xl
# import pandas as pd

# Разделение КП на архивные и не архивные по разным листам


print('start')
wb = xl.load_workbook('/files/Контейнерные площадки00.xlsx')
print('load_workbook done')

wb.create_sheet('Не архивные')
wb.create_sheet('Архивные')
print('create_sheets done')

sh = wb['Контейнерные площадки']
non_archive_sh = wb['Не архивные']
archive_sh = wb['Архивные']

archive_column = None

for col in sh.iter_cols(min_row=1, max_row=1):
    for cell in col:
        if cell.value == 'Архивная':
            archive_column = cell.column
            break

print('column find done')

if archive_column:
    for i in range(1, sh.max_column + 1):
        non_archive_sh.cell(1, i).value = sh.cell(1, i).value
        archive_sh.cell(1, i).value = sh.cell(1, i).value
    print('шапка done')

    counter = 0
    for row in sh.iter_rows(min_row=2, min_col=1):
        if counter % 500 == 0:
            print(counter)

        data = []
        for cell in row:
            data.append(cell.value)

        if row[archive_column - 1].value == 'Нет':
            non_archive_sh.append(data)
        elif row[archive_column - 1].value == 'Да':
            archive_sh.append(data)

        counter += 1

    print('for ended')
    wb.save('/files/Контейнерные площадки000.xlsx')
    print('Выполнено')
else:
    print('Столбец archive_column не найден')

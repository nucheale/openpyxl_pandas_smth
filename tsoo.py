import openpyxl as xl
from openpyxl import Workbook
from datetime import datetime
import os
import sys
# import pandas as pd

# Отчет по терсхеме

print('start')

files_dir = 'files/disp/'
files = os.listdir(files_dir)
if not files:
    sys.exit(1)

new_wb = Workbook()
new_sh = new_wb.active

skip_header = False
for wb in files:
    if wb[-4:] == 'xlsx':
        temp_wb = xl.load_workbook(f'{files_dir}{wb}')
        temp_sh = temp_wb.active
        for i, row in enumerate(temp_sh.iter_rows(min_row=4, values_only=True), start=4):
            if skip_header and i == 4:
                continue
            new_sh.append(row)
        skip_header = True
        temp_wb.close()
print('1')

rows_to_delete = []
schedules_for_delete = ['субботник', 'под загрузку', 'под погрузку', 'установка']
for row in new_sh.iter_rows(min_row=0):
    if '789' in (str(row[2].value.lower()))[:3]:
        rows_to_delete.append(row[0].row)
    if 'сигнальный' in row[3].value.lower() or 'вкп' in row[3].value.lower():
        rows_to_delete.append(row[0].row)
    if 'помешочный сбор' in str(row[6].value).lower():
        rows_to_delete.append(row[0].row)
    if any(schedule in str(row[14].value).lower() for schedule in schedules_for_delete):
        rows_to_delete.append(row[0].row)
    if 'контейнер выходного дня' in str(row[22].value).lower():
        rows_to_delete.append(row[0].row)
print('2')
rows_to_delete = list(set(rows_to_delete))
print(f'2 \n len: {len(rows_to_delete)}')

new_sh_2 = new_wb.create_sheet('Итог', 2)
for row in new_sh.iter_rows(min_row=0):
    if row[0].row not in rows_to_delete:
        row_with_values = [cell.value for cell in row]
        new_sh_2.append(row_with_values)
print('3')

new_wb.remove(new_sh)

new_wb.save(f'{files_dir}Итог/Реестр КП {datetime.now().strftime("%d.%m.%Y %H_%M")}.xlsx')
new_wb.close()

print('4')

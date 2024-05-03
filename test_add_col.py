import openpyxl as xl

# Найти столбец Код, после него вставить столбец Код КП новый, в него записать значение из столбца Код без "_000",
# перевести в число если возможно

wb = xl.load_workbook('/files/Контейнерные площадки0.xlsx')
sh = wb['Контейнерные площадки']

id_column = None

for col in sh.iter_cols(min_row=1, max_row=1):
    for cell in col:
        if cell.value == 'Код':
            id_column = cell.column
            break

if id_column:
    sh.insert_cols(id_column + 1)
    sh.cell(1, (id_column + 1), 'Код КП новый')
    for row in sh.iter_rows(min_row=2, min_col=id_column, max_col=id_column + 1):
        if '_000' in row[0].value:
            try:
                new_id = int(str(row[0].value).replace('_000', ''))
                row[1].value = new_id
            except ValueError:
                row[1].value = row[0].value
                pass

    wb.save('/files/Контейнерные площадки00.xlsx')
    print('Выполнено')
else:
    print('Столбец Код не найден')
